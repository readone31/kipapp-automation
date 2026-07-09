"""
spreadsheet_reader.py
Reads the input spreadsheet (Google Sheets public export or local .xlsx/.csv).
Expected columns (case-insensitive, flexible):
  Jenis Kinerja   | Rencana Kinerja | Kegiatan | Progres | Capaian |
  Data Dukung     | Tanggal Mulai   | Tanggal Selesai | Lama Hari

PENTING — Penanganan sel berformula:
Beberapa kolom pada spreadsheet sumber (mis. Progres, Capaian, Lama Hari)
kerap diisi dengan rumus (mis. "=B2-A2", "=AVERAGEIF(...)") dan bukan nilai
mentah. Modul ini SELALU berusaha mengambil HASIL PERHITUNGAN rumus
tersebut (bukan teks rumusnya) sebelum data diserahkan ke automation.py,
supaya string seperti "=B2-A2" tidak pernah ikut diketik ke form KipApp.
Lihat `_looks_like_formula()` dan `_recalculate_xlsx_with_libreoffice()`.
"""

import re
import io
import csv
import shutil
import subprocess
import tempfile
import urllib.request
from pathlib import Path
from typing import Any

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ── Alias kolom → kunci internal ─────────────────────────────────────────────
# Semua entri dalam huruf kecil tanpa spasi berlebih.
# Tambahkan alias baru di sini jika nama kolom di spreadsheet berbeda.
_COL_ALIASES = {
    # Jenis Kinerja (Utama / Tambahan) — menentukan alur Add di KipApp:
    # "Utama"    → menu Perencanaan Kinerja > Periodik > Add
    # "Tambahan" → menu Rencana Kinerja > Periode SKP > Add > Jenis Kinerja=Tambahan
    "jenis kinerja":            "jenis_kinerja",
    "jenis_kinerja":            "jenis_kinerja",
    "jeniskinerja":             "jenis_kinerja",
    "jenis":                    "jenis_kinerja",
    "kategori kinerja":         "jenis_kinerja",

    # Rencana Kinerja
    "rencana kinerja":          "rencana_kinerja",
    "rencana_kinerja":          "rencana_kinerja",
    "rencanakinerja":           "rencana_kinerja",
    "rencana":                  "rencana_kinerja",

    # Kegiatan
    "kegiatan":                 "kegiatan",

    # Progres
    "progres":                  "progres",
    "progress":                 "progres",
    "persentase progres":       "progres",
    "persentase_progres":       "progres",
    "% progres":                "progres",

    # Capaian
    "capaian":                  "capaian",
    "nilai capaian":            "capaian",
    "nilai_capaian":            "capaian",

    # Data Dukung
    "data dukung":              "data_dukung",
    "data_dukung":              "data_dukung",
    "datadukung":               "data_dukung",
    "bukti dukung":             "data_dukung",
    "bukti_dukung":             "data_dukung",

    # Tanggal Mulai
    "tanggal mulai":            "tanggal_mulai",
    "tanggal_mulai":            "tanggal_mulai",
    "tgl mulai":                "tanggal_mulai",
    "tgl_mulai":                "tanggal_mulai",
    "mulai":                    "tanggal_mulai",
    "start date":               "tanggal_mulai",
    "start":                    "tanggal_mulai",

    # Tanggal Selesai
    "tanggal selesai":          "tanggal_selesai",
    "tanggal_selesai":          "tanggal_selesai",
    "tgl selesai":              "tanggal_selesai",
    "tgl_selesai":              "tanggal_selesai",
    "selesai":                  "tanggal_selesai",
    "end date":                 "tanggal_selesai",
    "end":                      "tanggal_selesai",

    # Lama Hari
    "lama hari":                "lama_hari",
    "lama_hari":                "lama_hari",
    "lamahari":                 "lama_hari",
    "durasi":                   "lama_hari",
    "durasi hari":              "lama_hari",
    "jumlah hari":              "lama_hari",
    "jumlah_hari":              "lama_hari",
}

INTERNAL_KEYS = {
    "jenis_kinerja", "rencana_kinerja", "kegiatan", "progres", "capaian",
    "data_dukung", "tanggal_mulai", "tanggal_selesai", "lama_hari",
}


def _normalize_header(raw: str) -> str:
    """Normalkan nama kolom: lowercase, strip whitespace, cari alias."""
    cleaned = raw.strip().lower()
    # Hapus karakter BOM atau non-printable
    cleaned = re.sub(r'[\x00-\x1f\x7f-\x9f\ufeff]', '', cleaned)
    # Ganti multiple whitespace jadi satu spasi
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    return _COL_ALIASES.get(cleaned, cleaned)


def _google_sheets_to_csv_url(url: str) -> str | None:
    m = re.search(r"spreadsheets/d/([a-zA-Z0-9_-]+)", url)
    if not m:
        return None
    sheet_id = m.group(1)
    gid_m = re.search(r"[?&#]gid=(\d+)", url)
    gid = gid_m.group(1) if gid_m else "0"
    return (
        f"https://docs.google.com/spreadsheets/d/{sheet_id}"
        f"/export?format=csv&gid={gid}"
    )


def _rows_from_csv_text(text: str, log_fn=None) -> list[dict[str, Any]]:
    def _log(level, msg):
        if log_fn:
            log_fn(level, msg)

    reader = csv.DictReader(io.StringIO(text))
    normalized_rows = []
    formula_cells_blanked = 0
    for row in reader:
        normalized = {}
        for k, v in row.items():
            key = _normalize_header(k)
            val = v if v is not None else ""
            if _looks_like_formula(val):
                # Google Sheets CSV export normalnya sudah berupa hasil
                # perhitungan, tapi jika sel diformat sebagai teks murni
                # rumus bisa lolos apa adanya — jangan sampai ini terkirim
                # ke KipApp.
                formula_cells_blanked += 1
                val = ""
            normalized[key] = val
        normalized_rows.append(normalized)

    if formula_cells_blanked:
        _log("warning",
             f"{formula_cells_blanked} sel dari CSV masih berupa teks rumus "
             "mentah dan DIKOSONGKAN demi keamanan. Periksa format sel di "
             "spreadsheet sumber.")
    return normalized_rows


def _looks_like_formula(value: Any) -> bool:
    """Deteksi apakah sebuah nilai sel sebenarnya adalah teks rumus
    (mis. "=B2-A2", "=AVERAGEIF(range,kriteria,sum_range)") yang gagal
    terhitung — bukan hasil akhirnya. Ditandai dengan awalan '='."""
    if value is None:
        return False
    s = str(value).strip()
    return s.startswith("=") and len(s) > 1


def _recalculate_xlsx_with_libreoffice(path: str, log_fn=None) -> str | None:
    """Paksa LibreOffice (headless) membuka & menyimpan ulang file .xlsx.
    Ini membuat semua rumus dihitung ulang dan nilai HASIL-nya di-cache ke
    dalam file, sehingga pembacaan berikutnya dengan
    `openpyxl.load_workbook(data_only=True)` mendapat angka/teks hasil,
    bukan None atau teks rumus mentah.

    Mengembalikan path file hasil rekalkulasi (di folder temp), atau None
    jika LibreOffice tidak tersedia di komputer ini (fallback ditangani
    oleh pemanggil).
    """
    def _log(level, msg):
        if log_fn:
            log_fn(level, msg)

    soffice = (shutil.which("soffice") or shutil.which("libreoffice")
               or shutil.which("soffice.exe"))
    if not soffice:
        return None

    try:
        out_dir = tempfile.mkdtemp(prefix="kipapp_recalc_")
        _log("info", "Menghitung ulang rumus di spreadsheet via LibreOffice (headless)…")
        subprocess.run(
            [soffice, "--headless", "--norestore", "--calc",
             "--convert-to", "xlsx", "--outdir", out_dir, path],
            timeout=60, check=True,
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
        )
        recalculated = Path(out_dir) / (Path(path).stem + ".xlsx")
        if recalculated.exists():
            _log("success", "Rumus berhasil dihitung ulang, nilai hasil siap dipakai.")
            return str(recalculated)
    except Exception as e:
        _log("warning", f"Rekalkulasi via LibreOffice gagal ({e}) — lanjut dengan nilai cache.")
    return None


def _rows_from_xlsx(path: str, log_fn=None) -> list[dict[str, Any]]:
    """Baca .xlsx dan pastikan yang dikembalikan adalah NILAI HASIL rumus,
    bukan teks rumus itu sendiri.

    Strategi:
    1. Baca dengan data_only=True → openpyxl mengembalikan nilai HASIL
       terakhir yang di-cache Excel/Sheets saat file terakhir disimpan.
    2. Jika ada sel yang masih kosong padahal selnya sebetulnya berformula
       (cache belum pernah dihitung — umum terjadi kalau file dibuat/diedit
       lewat skrip tanpa pernah dibuka di Excel/Sheets), coba paksa
       LibreOffice menghitung ulang lalu baca ulang hasilnya.
    3. Sebagai jaring pengaman terakhir: kalau nilai yang terbaca TETAP
       berupa teks yang diawali '=' (rumus mentah), nilai itu DIKOSONGKAN
       (bukan dikirim apa adanya) dan dicatat sebagai warning — supaya
       automation.py tidak pernah mengetik rumus ke KipApp.
    """
    def _log(level, msg):
        if log_fn:
            log_fn(level, msg)

    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl not installed — cannot read .xlsx files.")

    wb_formulas = openpyxl.load_workbook(path, data_only=False)
    ws_formulas = wb_formulas.active

    wb_values = openpyxl.load_workbook(path, data_only=True)
    ws_values = wb_values.active

    needs_recalc = False
    for row_f in ws_formulas.iter_rows():
        for cell in row_f:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cached = ws_values.cell(row=cell.row, column=cell.column).value
                if cached is None or _looks_like_formula(cached):
                    needs_recalc = True
                    break
        if needs_recalc:
            break

    if needs_recalc:
        recalced_path = _recalculate_xlsx_with_libreoffice(path, log_fn=log_fn)
        if recalced_path:
            wb_values = openpyxl.load_workbook(recalced_path, data_only=True)
            ws_values = wb_values.active
        else:
            _log("warning",
                 "Beberapa sel berformula belum memiliki nilai hasil ter-cache "
                 "dan LibreOffice tidak tersedia untuk menghitung ulang. Buka & "
                 "simpan ulang file di Excel/Google Sheets sebelum dipakai, "
                 "atau sel tersebut akan dikosongkan demi keamanan.")

    rows_raw = list(ws_values.iter_rows(values_only=True))
    if not rows_raw:
        return []

    headers = [
        _normalize_header(str(h) if h is not None else "")
        for h in rows_raw[0]
    ]

    result: list[dict[str, Any]] = []
    formula_cells_blanked = 0
    for row_idx, row in enumerate(rows_raw[1:], start=2):
        d: dict[str, Any] = {}
        for i, header in enumerate(headers):
            raw_val = row[i] if i < len(row) else None
            text_val = "" if raw_val is None else str(raw_val).strip()
            if _looks_like_formula(text_val):
                # Jaring pengaman: JANGAN PERNAH kirim teks rumus ke KipApp.
                formula_cells_blanked += 1
                text_val = ""
            d[header] = text_val
        result.append(d)

    if formula_cells_blanked:
        _log("warning",
             f"{formula_cells_blanked} sel masih berupa teks rumus mentah dan "
             "DIKOSONGKAN demi keamanan (bukan dikirim ke KipApp). Simpan ulang "
             "file tersebut di Excel/Google Sheets agar nilainya ter-cache.")

    return result


def load_spreadsheet(url_or_path: str,
                     log_fn=None) -> list[dict[str, Any]]:
    """
    Load spreadsheet rows. Returns list of dicts keyed by internal names.
    Logs diagnostik kolom agar mudah debug jika ada kolom yang tidak terbaca.
    """
    def _log(level, msg):
        if log_fn:
            log_fn(level, msg)

    url_or_path = url_or_path.strip()

    # ── Baca sumber data ──────────────────────────────────────────────────
    if "docs.google.com/spreadsheets" in url_or_path:
        csv_url = _google_sheets_to_csv_url(url_or_path)
        if not csv_url:
            raise ValueError("URL Google Sheets tidak valid.")
        _log("info", "Mengunduh spreadsheet dari Google Sheets…")
        with urllib.request.urlopen(csv_url, timeout=30) as resp:
            text = resp.read().decode("utf-8-sig")
        rows = _rows_from_csv_text(text, log_fn=log_fn)

    elif url_or_path.lower().endswith((".xlsx", ".xlsm")):
        _log("info", f"Membaca file Excel: {url_or_path}")
        rows = _rows_from_xlsx(url_or_path, log_fn=log_fn)

    elif url_or_path.lower().endswith(".csv") or Path(url_or_path).exists():
        _log("info", f"Membaca file CSV: {url_or_path}")
        with open(url_or_path, encoding="utf-8-sig") as f:
            rows = _rows_from_csv_text(f.read(), log_fn=log_fn)

    else:
        _log("info", f"Mencoba mengunduh sebagai CSV: {url_or_path}")
        with urllib.request.urlopen(url_or_path, timeout=30) as resp:
            text = resp.read().decode("utf-8-sig")
        rows = _rows_from_csv_text(text, log_fn=log_fn)

    # ── Filter baris kosong ───────────────────────────────────────────────
    rows = [r for r in rows if any(
        v.strip() for v in r.values() if isinstance(v, str)
    )]

    # ── Diagnostik kolom ─────────────────────────────────────────────────
    if rows:
        found_keys   = set(rows[0].keys())
        mapped_keys  = found_keys & INTERNAL_KEYS
        missing_keys = INTERNAL_KEYS - found_keys

        _log("info", f"Kolom terbaca ({len(found_keys)}): {sorted(found_keys)}")
        _log("info", f"Kolom berhasil dipetakan: {sorted(mapped_keys)}")

        if missing_keys:
            _log("warning",
                 f"Kolom TIDAK ditemukan (akan dibaca kosong): {sorted(missing_keys)}. "
                 "Periksa nama kolom di spreadsheet Anda.")
        else:
            _log("success", "Semua kolom wajib berhasil dipetakan.")

        # Preview baris pertama untuk debug
        preview = {k: v for k, v in rows[0].items() if k in INTERNAL_KEYS}
        _log("info", f"Preview baris 1: {preview}")

    _log("info", f"Spreadsheet dimuat: {len(rows)} baris data.")
    return rows
